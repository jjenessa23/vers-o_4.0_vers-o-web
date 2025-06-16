import os
import streamlit as st
import pandas as pd
from datetime import datetime
import re
import io # Para manipulação de arquivos em memória
import openpyxl # Para gerar e ler arquivos Excel
from openpyxl.styles import Font, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, A4, landscape # Importar landscape
from reportlab.lib.units import inch, mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image, Frame, PageTemplate, NextPageTemplate # Importar Frame, PageTemplate, e NextPageTemplate
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
import logging
import sqlite3 # Importar sqlite3 para verificar tipo de dado
from app_logic.utils import set_background_image, set_sidebar_background_image

# Importar funções do novo módulo de utilitários de banco de dados
from db_utils import get_declaracao_by_referencia, get_itens_by_declaracao_id, update_xml_item_erp_code, get_process_cost_data, save_process_cost_data

logger = logging.getLogger(__name__)

# --- Funções Auxiliares de Formatação ---
def _format_currency(value):
    """Formata um valor numérico como moeda BRL."""
    try:
        val = float(value)
        return f"R$ {val:,.2f}".replace('.', '#').replace(',', '.').replace('#', ',')
    except (ValueError, TypeError):
        return "R$ 0,00"

def _format_float(value, decimals=6, prefix=""):
    """Formata um valor numérico como float com número específico de casas decimais."""
    try:
        val = float(value)
        if val == 0:
            return f"{prefix}{0:.{decimals}f}".replace('.', '#').replace(',', '.').replace('#', ',')
        return f"{prefix}{val:,.{decimals}f}".replace('.', '#').replace(',', '.').replace('#', ',')
    except (ValueError, TypeError):
        return "N/A"

def _format_percent(value):
    """Formata um valor numérico como porcentagem."""
    try:
        val = float(value) * 100
        return f"{val:,.2f}%".replace('.', '#').replace(',', '.').replace('#', ',')
    except (ValueError, TypeError):
        return "0,00%"

def _format_weight_no_kg(value):
    """Formata um valor numérico como peso em KG."""
    try:
        val = float(value)
        return f"{val:,.3f} KG".replace('.', '#').replace(',', '.').replace('#', ',')
    except (ValueError, TypeError):
        return "0,000 KG"

def _format_int(value):
    """Formata um valor numérico como inteiro."""
    try:
        val = int(value)
        return str(val)
    except (ValueError, TypeError):
        return "N/A"

def _format_int_no_float(value):
    """Formata um valor numérico que pode ser float para inteiro."""
    try:
        return str(int(float(value)))
    except (ValueError, TypeError):
        return "0"

def _format_ncm(ncm_value):
    """Formata o NCM com pontos."""
    if ncm_value and isinstance(ncm_value, str) and len(ncm_value) == 8:
        return f"{ncm_value[0:4]}.{ncm_value[4:6]}.{ncm_value[6:8]}"
    return ncm_value

def _format_di_number(di_number):
    """Formata o número da DI."""
    if di_number and isinstance(di_number, str) and len(di_number) == 10:
        return f"{di_number[0:2]}/{di_number[2:9]}-{di_number[9]}"
    return di_number

# Adicione após as outras funções auxiliares de formatação
def _clean_number(x):
    """Limpa string numérica removendo KG e convertendo para formato float"""
    try:
        # Remove KG e espaços
        x = str(x).replace(' KG', '').strip()
        # Remove todos os pontos exceto o último (para números como 3.625.909)
        if x.count('.') > 1:
            last_dot_index = x.rindex('.')
            x = x[:last_dot_index].replace('.', '') + x[last_dot_index:]
        # Substitui vírgula por ponto
        x = x.replace(',', '.')
        return float(x)
    except (ValueError, AttributeError):
        return 0.0

def _clean_quantity(x):
    """Limpa string numérica tratando separador de milhar corretamente"""
    try:
        # Remove espaços
        x = str(x).strip()
        # Para números como 2.000, mantém os pontos como separador de milhar
        if ',' not in x:  # Se não tem vírgula, ponto é separador de milhar
            # Apenas remove os pontos se não houver vírgula (considera ponto como milhar)
            return float(x.replace('.', ''))
        # Se tem vírgula, trata ponto como milhar e vírgula como decimal
        x = x.replace('.', '').replace(',', '.')
        return float(x)
    except (ValueError, AttributeError):
        return 0.0

# --- Função de Cálculo Principal (Adaptada do seu código) ---
def perform_calculations(di_data, itens_data, expense_inputs, contracts_df):
    """Realiza todos os cálculos de custo do processo e itens."""
    if not di_data:
        return {}, {}, {}, pd.DataFrame(), 0.0, 0.0

    # Desempacota os dados da DI
    (id_db, numero_di, data_registro_db, valor_total_reais_xml,
     arquivo_origem, data_importacao, informacao_complementar,
     vmle_declaracao, frete_declaracao, seguro_declaracao, vmld_declaracao,
     ipi_total_declaracao, pis_pasep_total_declaracao, cofins_total_declaracao, icms_sc,
     taxa_cambial_usd_declaracao, taxa_siscomex_total_declaracao, numero_invoice,
     peso_bruto_total, peso_liquido_total, cnpj_importador, importador_nome,
     recinto, embalagem, quantidade_volumes_total, acrescimo_total_declaracao,
     imposto_importacao_total_declaracao, armazenagem_db, frete_nacional_db) = di_data

    # Obter valores dos campos editáveis de despesas
    afrmm_input = expense_inputs['afrmm']
    siscoserv_input = expense_inputs['siscoserv']
    descarregamento_input = expense_inputs['descarregamento']
    taxas_destino_input = expense_inputs['taxas_destino']
    multa_input = expense_inputs['multa']

    # Cálculo de Despesas Operacionais (Processo)
    envio_docs_fixo = 0.00
    honorario_despachante_fixo = 1000.00

    total_despesas_operacionais = (
        afrmm_input + (armazenagem_db if armazenagem_db is not None else 0.0) + envio_docs_fixo +
        (frete_nacional_db if frete_nacional_db is not None else 0.0) + honorario_despachante_fixo +
        (taxa_siscomex_total_declaracao if taxa_siscomex_total_declaracao is not None else 0.0) + siscoserv_input +
        descarregamento_input + taxas_destino_input + multa_input
    )

    # Cálculo dos Contratos de Câmbio
    soma_contratos_reais = 0.0
    soma_contratos_usd = 0.0
    for index, row in contracts_df.iterrows():
        try:
            dolar_val = row['Dólar']
            valor_contrato_usd_input = row['Valor (US$)']

            if dolar_val > 0 and valor_contrato_usd_input > 0:
                soma_contratos_reais += (dolar_val * valor_contrato_usd_input)
                soma_contratos_usd += valor_contrato_usd_input
        except (ValueError, TypeError):
            pass

    if acrescimo_total_declaracao is not None and taxa_cambial_usd_declaracao is not None and taxa_cambial_usd_declaracao > 0:
        soma_contratos_usd += (acrescimo_total_declaracao / taxa_cambial_usd_declaracao)

    vmle_declaracao_safe = vmle_declaracao if vmle_declaracao is not None else 0.0
    acrescimo_total_declaracao_safe = acrescimo_total_declaracao if acrescimo_total_declaracao is not None else 0.0
    variacao_cambial_total = soma_contratos_reais - (vmle_declaracao_safe - acrescimo_total_declaracao_safe)

    # Totais do Processo
    cambio_di_para_usd = taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao is not None and taxa_cambial_usd_declaracao > 0 else None

    process_totals = {
        "Taxa Cambial": _format_float(taxa_cambial_usd_declaracao, 6) if taxa_cambial_usd_declaracao is not None else "N/A",
        "VMLE (R$)": _format_currency(vmle_declaracao) if vmle_declaracao is not None else "R$ 0,00",
        "VMLE (US$)": _format_float(vmle_declaracao / cambio_di_para_usd, 2, prefix="US$ ") if vmle_declaracao is not None and cambio_di_para_usd else "US$ 0,00",
        "Frete (R$)": _format_currency(frete_declaracao) if frete_declaracao is not None else "R$ 0,00",
        "Frete (US$)": _format_float(frete_declaracao / cambio_di_para_usd, 2, prefix="US$ ") if frete_declaracao is not None and cambio_di_para_usd else "US$ 0,00",
        "Seguro (R$)": _format_currency(seguro_declaracao) if seguro_declaracao is not None else "R$ 0,00",
        "Seguro (US$)": _format_float(seguro_declaracao / cambio_di_para_usd, 2, prefix="US$ ") if seguro_declaracao is not None and cambio_di_para_usd else "US$ 0,00",
        "VMLD (CIF) (R$)": _format_currency(vmld_declaracao) if vmld_declaracao is not None else "R$ 0,00",
        "VMLD (CIF) (US$)": _format_float(vmld_declaracao / cambio_di_para_usd, 2, prefix="US$ ") if vmld_declaracao is not None and cambio_di_para_usd else "US$ 0,00",
        "Acréscimo (R$)": _format_currency(acrescimo_total_declaracao) if acrescimo_total_declaracao is not None else "R$ 0,00",
        "Acréscimo (US$)": _format_float(acrescimo_total_declaracao / cambio_di_para_usd, 2, prefix="US$ ") if acrescimo_total_declaracao is not None and cambio_di_para_usd else "US$ 0,00",
        "Peso Total (KG)": _format_weight_no_kg(peso_liquido_total) if peso_liquido_total is not None else "0,000 KG",
        "SISCOMEX": _format_currency(taxa_siscomex_total_declaracao) if taxa_siscomex_total_declaracao is not None else "R$ 0,00",
        "Despesas Operacionais": _format_currency(total_despesas_operacionais),
    }

    # Impostos
    taxes_data = {
        "II": _format_currency(imposto_importacao_total_declaracao) if imposto_importacao_total_declaracao is not None else "R$ 0,00",
        "IPI": _format_currency(ipi_total_declaracao) if ipi_total_declaracao is not None else "R$ 0,00",
        "PIS": _format_currency(pis_pasep_total_declaracao) if pis_pasep_total_declaracao is not None else "R$ 0,00",
        "COFINS": _format_currency(cofins_total_declaracao) if cofins_total_declaracao is not None else "R$ 0,00",
    }

    # Despesas (para exibição)
    expenses_display = {
        "AFRMM": _format_currency(afrmm_input),
        "ARMAZENAGEM": _format_currency(armazenagem_db),
        "ENVIO DE DOCS": _format_currency(envio_docs_fixo),
        "FRETE NACIONAL": _format_currency(frete_nacional_db) if frete_nacional_db is not None else "R$ 0,00",
        "HONORÁRIO DESPACHANTE": _format_currency(honorario_despachante_fixo),
        "SISCOMEX": _format_currency(taxa_siscomex_total_declaracao),
        "SISCOSERV": _format_currency(siscoserv_input),
        "DESCARREGAMENTO": _format_currency(descarregamento_input),
        "TAXAS DESTINO": _format_currency(taxas_destino_input),
        "MULTA": _format_currency(multa_input),
        "TOTAL": _format_currency(total_despesas_operacionais),
    }

    total_para_nf = (
        (vmle_declaracao if vmle_declaracao is not None else 0.0) +
        (frete_declaracao if frete_declaracao is not None else 0.0) +
        (seguro_declaracao if seguro_declaracao is not None else 0.0) +
        (imposto_importacao_total_declaracao if imposto_importacao_total_declaracao is not None else 0.0) +
        (ipi_total_declaracao if ipi_total_declaracao is not None else 0.0) +
        (pis_pasep_total_declaracao if pis_pasep_total_declaracao is not None else 0.0) +
        (cofins_total_declaracao if cofins_total_declaracao is not None else 0.0) +
        total_despesas_operacionais
    )
    expenses_display["TOTAL PARA NF"] = _format_currency(total_para_nf)

    # Diferença Contratos
    vmle_declaracao_usd = vmle_declaracao_safe / taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao > 0 else 0.0
    diferenca_contratos_usd = soma_contratos_usd - vmle_declaracao_usd

    # Cálculos e População da Tabela de Itens
    itens_df_data = []
    total_peso_liquido_itens_di = sum(item[9] for item in itens_data if item[9] is not None)
    total_valor_fob_brl_itens_di = sum(item[8] for item in itens_data if item[8] is not None)
    total_quantidade_itens_di = 0
    for item in itens_data:
        if item[5] is not None:
            try:
                # Use _clean_quantity para obter o valor numérico correto
                qty = _clean_quantity(item[5])
                # Correção: dividir a quantidade por 10 conforme solicitado
                qty = qty / 10.0
                total_quantidade_itens_di += qty
            except (ValueError, AttributeError):
                continue

    total_peso_liquido_itens_di = total_peso_liquido_itens_di if total_peso_liquido_itens_di > 0 else 1.0
    total_valor_fob_brl_itens_di = total_valor_fob_brl_itens_di if total_valor_fob_brl_itens_di > 0 else 1.0
    total_quantidade_itens_di = total_quantidade_itens_di if total_quantidade_itens_di > 0 else 1.0

    vmld_declaracao_para_rateio = vmld_declaracao if vmld_declaracao is not None and vmld_declaracao > 0 else 1.0

    # Calculate total VLME for all items (needed for "Seguro do item" calculation)
    total_vlme_brl_itens_di_calc = 0.0
    for item_data in itens_data:
        # Garante que item_data seja uma tupla/lista e não sqlite3.Row
        if isinstance(item_data, sqlite3.Row):
            item_data = tuple(item_data)
        
        qty_original = item_data[5] if item_data[5] is not None else 0
        qty = _clean_quantity(qty_original) / 10.0 # Aplicar a divisão por 10 aqui também para cálculos
        
        val_item_fob_brl = item_data[8] if item_data[8] is not None else 0.0
        peso_liquido_item_from_db = item_data[9] if item_data[9] is not None else 0.0
        acrescimo_rateado_item_brl_calc = (acrescimo_total_declaracao if acrescimo_total_declaracao is not None else 0.0) / total_peso_liquido_itens_di * peso_liquido_item_from_db if total_peso_liquido_itens_di > 0 else 0.0
        vlme_brl_item_calc = val_item_fob_brl + acrescimo_rateado_item_brl_calc
        total_vlme_brl_itens_di_calc += vlme_brl_item_calc
    total_vlme_brl_itens_di_calc = total_vlme_brl_itens_di_calc if total_vlme_brl_itens_di_calc > 0 else 1.0

    fatores_por_adicao = {}

    for item_data in itens_data:
        # Ensure item_data is a tuple/list, not sqlite3.Row
        if isinstance(item_data, sqlite3.Row):
            item_data = tuple(item_data)

        (item_id, decl_id, num_adicao, num_item_seq, desc_mercadoria, qty_original, unit_medida, # Renomeado qty para qty_original
         val_unit_fob_usd, val_item_fob_brl, peso_liquido_item_from_db, ncm_item, sku_item,
         custo_unit_di_usd, ii_perc_item, ipi_perc_item, pis_perc_item, cofins_perc_item, icms_perc_item,
         codigo_erp_do_db) = item_data

        # Use _clean_quantity para garantir que a quantidade seja um número correto
        qty = _clean_quantity(qty_original) if qty_original is not None else 0
        # Correção: dividir a quantidade por 10 para todos os cálculos
        qty = qty / 10.0

        val_item_fob_brl = val_item_fob_brl if val_item_fob_brl is not None else 0.0
        peso_liquido_item_from_db = peso_liquido_item_from_db if peso_liquido_item_from_db is not None else 0.0
        custo_unit_di_usd = custo_unit_di_usd if custo_unit_di_usd is not None else 0.0
        taxa_cambial_usd_proc = taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao is not None else 0.0

        peso_liquido_item_rateado = peso_liquido_item_from_db
        frete_rateado_item = (frete_declaracao / total_peso_liquido_itens_di) * peso_liquido_item_rateado if total_peso_liquido_itens_di > 0 else 0.0
        acrescimo_rateado_item_brl = (acrescimo_total_declaracao if acrescimo_total_declaracao is not None else 0.0) / total_peso_liquido_itens_di * peso_liquido_item_rateado if total_peso_liquido_itens_di > 0 else 0.0
        vlme_brl_item = val_item_fob_brl + acrescimo_rateado_item_brl
        seguro_rateado_item = (seguro_declaracao / total_vlme_brl_itens_di_calc) * vlme_brl_item if total_vlme_brl_itens_di_calc > 0 else 0.0
        vlmd_brl_item = vlme_brl_item + frete_rateado_item + seguro_rateado_item
        cif_item_total = vlmd_brl_item
        cif_unitario_item = cif_item_total / qty if qty > 0 else 0.0

        ii_perc_item_safe = ii_perc_item if ii_perc_item is not None else 0.0
        ipi_perc_item_safe = ipi_perc_item if ipi_perc_item is not None else 0.0
        pis_perc_item_safe = pis_perc_item if pis_perc_item is not None else 0.0
        cofins_perc_item_safe = cofins_perc_item if cofins_perc_item is not None else 0.0
        icms_perc_item_safe = icms_perc_item if icms_perc_item is not None else 0.0


        ii_item_val_brl = vlmd_brl_item * ii_perc_item_safe
        ipi_item_val_brl = (vlmd_brl_item + ii_item_val_brl) * ipi_perc_item_safe
        pis_item_val_brl = vlmd_brl_item * pis_perc_item_safe
        cofins_item_val_brl = vlmd_brl_item * cofins_perc_item_safe
        icms_item_val = cif_unitario_item * icms_perc_item_safe

        despesas_rateada_item = (total_despesas_operacionais / vmld_declaracao_para_rateio) * vlmd_brl_item if vmld_declaracao_para_rateio > 0 else 0.0
        total_de_despesas_item = vlmd_brl_item + ii_item_val_brl + ipi_item_val_brl + pis_item_val_brl + cofins_item_val_brl + despesas_rateada_item
        total_unitario_item = total_de_despesas_item / qty if qty > 0 else 0.0
        item_variacao_cambial = variacao_cambial_total / total_quantidade_itens_di if total_quantidade_itens_di > 0 else 0.0
        total_unitario_com_variacao = total_unitario_item + item_variacao_cambial
        fator_internacao = total_unitario_com_variacao / (custo_unit_di_usd * taxa_cambial_usd_proc) if (custo_unit_di_usd * taxa_cambial_usd_proc) > 0 else 0.0

        if num_adicao not in fatores_por_adicao:
            fatores_por_adicao[num_adicao] = []
        fatores_por_adicao[num_adicao].append(fator_internacao)

        # Ajuste para extrair o SKU da descrição: Captura tudo até o primeiro " - " (espaço, traço, espaço)
        extracted_sku = sku_item # Valor padrão
        if desc_mercadoria:
            match = re.match(r'^(.*?)\s-\s', desc_mercadoria) # Expressão regular ajustada para " - " com espaços
            if match:
                extracted_sku = match.group(1).strip()
            else:
                extracted_sku = sku_item if sku_item else "N/A"
        else:
            extracted_sku = sku_item if sku_item else "N/A"


        itens_df_data.append({
            "ID": item_id,
            "Código ERP": st.session_state.item_erp_codes.get(item_id, codigo_erp_do_db if codigo_erp_do_db else ""), # Recebe o código ERP do banco
            "NCM": _format_ncm(ncm_item),
            "SKU": extracted_sku, # Usando o SKU extraído
            "Descrição": desc_mercadoria if desc_mercadoria else "N/A", # Mantém a descrição original
            "Quantidade": _format_int(qty), # Usando _format_int para exibir corretamente
            "Peso Unitário": _format_weight_no_kg(peso_liquido_item_rateado),
            "CIF Unitário": _format_float(cif_unitario_item, 4, prefix="R$ "),
            "VLME (BRL)": _format_currency(vlme_brl_item),
            "VLMD (BRL)": _format_currency(vlmd_brl_item),
            "II (BRL)": _format_currency(ii_item_val_brl),
            "IPI (BRL)": _format_currency(ipi_item_val_brl),
            "PIS (BRL)": _format_currency(pis_item_val_brl),
            "COFINS (BRL)": _format_currency(cofins_item_val_brl),
            "II %": _format_percent(ii_perc_item),
            "IPI %": _format_percent(ipi_perc_item),
            "PIS %": _format_percent(pis_perc_item),
            "COFINS %": _format_percent(cofins_perc_item),
            "ICMS %": _format_percent(icms_perc_item),
            "Frete R$": _format_currency(frete_rateado_item),
            "Seguro R$": _format_currency(seguro_rateado_item),
            "Unitário US$ DI": _format_float(custo_unit_di_usd, 2),
            "Despesas Rateada": _format_currency(despesas_rateada_item),
            "Total de Despesas": _format_currency(total_de_despesas_item),
            "Total Unitário": _format_currency(total_unitario_item),
            "Variação Cambial": _format_currency(item_variacao_cambial),
            "Total Unitário com Variação": _format_currency(total_unitario_com_variacao),
            "Fator de Internação": _format_float(fator_internacao, 4),
            "Fator por Adição": "Calculando..." # Será preenchido no final
        })

    itens_df = pd.DataFrame(itens_df_data)

    # Calcular Fator por Adição e Fator Geral
    adicao_fator_medio = {}
    for adicao, fatores in fatores_por_adicao.items():
        if fatores:
            adicao_fator_medio[adicao] = sum(fatores) / len(fatores)
        else:
            adicao_fator_medio[adicao] = 0.0

    # Atualizar a coluna "Fator por Adição" no DataFrame
    for index, row in itens_df.iterrows():
        # Para o mock, vamos simplesmente aplicar o fator médio geral:
        if adicao_fator_medio:
            itens_df.loc[index, "Fator por Adição"] = _format_float(sum(adicao_fator_medio.values()) / len(adicao_fator_medio), 4)


    total_impostos_processo = (imposto_importacao_total_declaracao if imposto_importacao_total_declaracao is not None else 0.0) + \
                              (ipi_total_declaracao if ipi_total_declaracao is not None else 0.0) + \
                              (pis_pasep_total_declaracao if pis_pasep_total_declaracao is not None else 0.0) + \
                              (cofins_total_declaracao if cofins_total_declaracao is not None else 0.0)

    fator_geral_numerador = total_impostos_processo + \
                            (frete_declaracao if frete_declaracao is not None else 0.0) + \
                            total_despesas_operacionais + \
                            (seguro_declaracao if seguro_declaracao is not None else 0.0) + \
                            (acrescimo_total_declaracao if acrescimo_total_declaracao is not None else 0.0) + \
                            soma_contratos_reais

    fator_geral_denominador = (vmle_declaracao_safe - (acrescimo_total_declaracao if acrescimo_total_declaracao is not None else 0.0))
    if fator_geral_denominador == 0:
        fator_geral_denominador = 1.0
    fator_geral_total = fator_geral_numerador / fator_geral_denominador
    process_totals["Fator Geral"] = _format_float(fator_geral_total, 4)

    # Adicionar linha de total ao DataFrame de itens
    total_row_data = {col: "" for col in itens_df.columns}
    total_row_data["Código ERP"] = "TOTAL"
    total_row_data["Quantidade"] = _format_int(itens_df["Quantidade"].apply(_clean_quantity).sum()) # Usando _format_int
    total_row_data["Peso Unitário"] = _format_weight_no_kg(
        itens_df["Peso Unitário"].apply(_clean_number).sum()
    )
    # Para colunas formatadas como "R$ X,XX" ou "US$ X,XX", precisamos remover o prefixo e converter para float antes de somar
    cols_to_sum_currency = ["CIF Unitário", "VLME (BRL)", "VLMD (BRL)", "II (BRL)", "IPI (BRL)", "PIS (BRL)", "COFINS (BRL)",
                            "Frete R$", "Seguro R$", "Despesas Rateada", "Total de Despesas", "Total Unitário", "Variação Cambial", "Total Unitário com Variação"]
    for col in cols_to_sum_currency:
        total_row_data[col] = _format_currency(itens_df[col].apply(lambda x: float(str(x).replace('R$', '').replace('US$', '').replace('.', '').replace(',', '.').strip())).sum())

    total_row_data["Unitário US$ DI"] = _format_float(itens_df["Unitário US$ DI"].apply(lambda x: float(str(x).replace('US$', '').replace('.', '').replace(',', '.').strip())).sum(), 2, prefix="US$ ")

    # Para Fator de Internação e Fator por Adição, calcular média dos itens
    overall_fator_internacao = itens_df["Fator de Internação"].apply(lambda x: float(str(x).replace('.', '').replace(',', '.'))).mean() if not itens_df.empty else 0.0
    overall_fator_por_adicao = itens_df["Fator por Adição"].apply(lambda x: float(str(x).replace('.', '').replace(',', '.'))).mean() if not itens_df.empty else 0.0
    total_row_data["Fator de Internação"] = _format_float(overall_fator_internacao, 4)
    total_row_data["Fator por Adição"] = _format_float(overall_fator_por_adicao, 4)

    itens_df = pd.concat([itens_df, pd.DataFrame([total_row_data])], ignore_index=True)

    return process_totals, taxes_data, expenses_display, itens_df, soma_contratos_usd, diferenca_contratos_usd

# --- Funções de Geração de Arquivos ---
def _generate_excel_for_cadastro(di_data, itens_data, item_erp_codes):
    """Gera um arquivo Excel com os dados dos itens para solicitação de cadastro."""
    if not di_data or not itens_data:
        st.warning("Nenhum dado de DI ou itens carregado para gerar o Excel.")
        return None, None

    referencia_di = di_data[6] if di_data[6] else "SemReferencia"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{referencia_di} - Itens para solicitação de cadastro"

    # Ajustado para que "ID do Item" seja a última coluna no Excel
    headers = ["COD", "SKU", "Descrição", "NCM", "Referência", "ID do Item"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_idx, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header_text)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        cell.border = thin_border


    for item_data in itens_data:
        # Ensure item_data is a tuple or list, and access by index
        # If it's a sqlite3.Row object, convert to tuple/list first
        if isinstance(item_data, sqlite3.Row):
            item_data = tuple(item_data)

        (item_id, decl_id, num_adicao, num_item_seq, desc_mercadoria, qty, unit_medida,
         val_unit_fob_usd, val_item_calculado_fob_brl, peso_liquido_item, ncm_item, sku_item,
         custo_unit_di_usd, ii_perc_item, ipi_perc_item, pis_perc_item, cofins_perc_item, icms_perc_item,
         codigo_erp_do_db) = item_data

        display_desc_mercadoria = desc_mercadoria
        # Usando a mesma lógica de extração de SKU para o Excel
        extracted_sku = sku_item
        if desc_mercadoria:
            match = re.match(r'^(.*?)\s-\s', desc_mercadoria) # Expressão regular ajustada para " - " com espaços
            if match:
                extracted_sku = match.group(1).strip()
            else:
                extracted_sku = sku_item if sku_item else "N/A"
        else:
            extracted_sku = sku_item if sku_item else "N/A"
        
        formatted_ncm = _format_ncm(ncm_item)

        row_data = [
            item_erp_codes.get(item_id, ""), # COD (Código ERP atual)
            extracted_sku, # SKU (extraído ou original)
            display_desc_mercadoria,
            formatted_ncm,
            referencia_di, # Referência da DI
            item_id # ID do Item (última coluna)
        ]
        ws.append(row_data)

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_idx)
            cell.border = thin_border


    for column in ws.columns:
        max_length = 0
        column_name = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column_name].width = adjusted_width

    # Salvar em um buffer de memória
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer, f"{referencia_di}_Itens_Cadastro.xlsx"

def _import_excel_for_cadastro(uploaded_file, itens_data):
    """
    Importa um arquivo Excel para atualizar os Códigos ERP dos itens.
    Espera a estrutura: COD, ID do Item (para correlação).
    """
    if uploaded_file is None:
        st.warning("Nenhum arquivo para importar.")
        return 0

    try:
        # Usar pandas para ler o Excel
        df = pd.read_excel(uploaded_file)
        
        # Agora a correlação será feita pelo "ID do Item"
        if 'COD' not in df.columns or 'ID do Item' not in df.columns:
            st.error("O arquivo Excel deve conter as colunas 'COD' e 'ID do Item'.")
            return 0

        updates_count = 0
        # Criar um mapeamento de item_id para item_tuple para busca eficiente
        itens_data_map = {item[0]: item for item in itens_data}

        for _, row in df.iterrows():
            erp_code = row['COD']
            item_id_from_excel = row['ID do Item']

            if pd.isna(erp_code) or pd.isna(item_id_from_excel):
                continue
            
            # Converter item_id_from_excel para o mesmo tipo que o item_id do banco (int)
            try:
                item_id_from_excel = int(item_id_from_excel)
            except ValueError:
                st.warning(f"ID do Item '{item_id_from_excel}' no Excel não é um número válido. Pulando esta linha.")
                continue

            # Buscar o item_id diretamente no mapeamento
            if item_id_from_excel in itens_data_map:
                found_item_id = item_id_from_excel
                if update_xml_item_erp_code(found_item_id, str(erp_code).strip()):
                    st.session_state.item_erp_codes[found_item_id] = str(erp_code).strip()
                    updates_count += 1
            else:
                st.warning(f"ID do Item '{item_id_from_excel}' não encontrado nos dados carregados da DI. Pulando esta linha.")

        return updates_count

    except Exception as e:
        st.error(f"Ocorreu um erro ao importar o arquivo Excel: {e}")
        logger.exception("Erro ao importar Código ERP do Excel.")
        return 0

def _generate_process_report_pdf(di_data, itens_df_calculated, soma_contratos_usd, diferenca_contratos_usd):
    """Gera um relatório completo do processo em PDF."""
    if not di_data or itens_df_calculated.empty:
        st.warning("Nenhum dado de DI ou itens carregado para gerar o relatório.")
        return None, None

    referencia_processo = di_data[6] if di_data[6] else "SemReferencia"
    file_name = f"{referencia_processo}_Relatorio.pdf"

    buffer = io.BytesIO() 
    doc = SimpleDocTemplate(buffer) # Inicializa sem pagesize para adicionar templates

    # Define frames para o modo retrato
    # Margens: 1 inch de cada lado (topo, base, esq, dir)
    leftMargin, rightMargin, topMargin, bottomMargin = inch, inch, inch, inch

    frame_portrait = Frame(leftMargin, bottomMargin, 
                           A4[0] - leftMargin - rightMargin, # Largura útil
                           A4[1] - topMargin - bottomMargin, # Altura útil
                           id='portrait_frame')

    # Define frames para o modo paisagem
    # landscape(A4) é (altura_A4, largura_A4)
    # Correção: Definir todas as margens para paisagem explicitamente
    landscape_left_margin, landscape_bottom_margin = inch, inch
    landscape_right_margin, landscape_top_margin = inch, inch # Definir as margens
    landscape_width_usable = landscape(A4)[0] - landscape_left_margin - landscape_right_margin # Largura útil em paisagem
    landscape_height_usable = landscape(A4)[1] - landscape_top_margin - landscape_bottom_margin # Altura útil em paisagem

    frame_landscape = Frame(landscape_left_margin, landscape_bottom_margin, 
                            landscape_width_usable,
                            landscape_height_usable,
                            id='landscape_frame')

    # Define PageTemplates
    portrait_template = PageTemplate(id='PortraitPage', frames=[frame_portrait], pagesize=A4)
    landscape_template = PageTemplate(id='LandscapePage', frames=[frame_landscape], pagesize=landscape(A4))

    # Adiciona os templates ao documento
    doc.addPageTemplates([portrait_template, landscape_template])
    
    # Importante: story precisa ser uma lista
    story = []
    
    styles = getSampleStyleSheet()
    
    style_title = ParagraphStyle(name='TitleStyle', parent=styles['h1'], fontSize=16, alignment=TA_CENTER, spaceAfter=14)
    style_heading = ParagraphStyle(name='HeadingStyle', parent=styles['h2'], fontSize=12, spaceAfter=8, alignment=TA_LEFT)
    style_normal = styles['Normal']
    style_normal.fontSize = 10
    style_normal.leading = 12
    style_bold = ParagraphStyle(name='BoldStyle', parent=style_normal, fontName='Helvetica-Bold')

    # Conteúdo das primeiras páginas (modo retrato)
    story.append(Paragraph(f"Relatório do Processo de Importação - DI: {_format_di_number(di_data[1])}", style_title))
    story.append(Spacer(1, 0.2*inch))

    # --- Dados Gerais da DI ---
    story.append(Paragraph("Dados Gerais da Declaração de Importação:", style_heading))
    di_general_data = [
        ["Referência:", di_data[6]],
        ["Número DI:", _format_di_number(di_data[1])],
        ["Data DI:", datetime.strptime(di_data[2], "%Y-%m-%d").strftime("%d/%m/%Y") if di_data[2] else "N/A"],
        ["VMLE:", _format_currency(di_data[7])],
        ["Frete DI:", _format_currency(di_data[8])],
        ["Seguro DI:", _format_currency(di_data[9])],
        ["VMLD (CIF):", _format_currency(di_data[10])],
        ["Taxa Cambial (USD):", _format_float(di_data[15], 6)],
        ["Nº Invoice:", di_data[17]],
        ["Peso Bruto Total (KG):", _format_weight_no_kg(di_data[18])],
        ["Peso Líquido Total (KG):", _format_weight_no_kg(di_data[19])],
        ["CNPJ Importador:", di_data[20]],
        ["Importador Nome:", di_data[21]],
        ["Recinto:", di_data[22]],
        ["Embalagem:", di_data[23]],
        ["Quantidade Volumes:", _format_int(di_data[24])],
        ["Acréscimo:", _format_currency(di_data[25])],
        ["Imposto de Importação:", _format_currency(di_data[26])],
        ["Armazenagem (DB):", _format_currency(di_data[27])],
        ["Frete Nacional (DB):", _format_currency(di_data[28])],
    ]
    table_di_general = Table(di_general_data, colWidths=[2.5*inch, 5*inch])
    table_di_general.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'LEFT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(table_di_general)
    story.append(Spacer(1, 0.2*inch))

    # --- Totais do Processo ---
    story.append(Paragraph("Totais do Processo:", style_heading))
    process_totals_data = [
        ["Item", "Valor (R$)", "Valor (US$)"]
    ]
    # Use st.session_state.process_totals for current values
    process_total_items_for_pdf = [
        ("Taxa Cambial", "Taxa Cambial"),
        ("VMLE", "VMLE (R$)"),
        ("Frete", "Frete (R$)"),
        ("Seguro", "Seguro (R$)"),
        ("VMLD (CIF)", "VMLD (CIF) (R$)"),
        ("Acréscimo", "Acréscimo (R$)"),
        ("Peso Total (KG)", "Peso Total (KG)"),
        ("SISCOMEX", "SISCOMEX"),
        ("Despesas Operacionais", "Despesas Operacionais"),
        ("Fator Geral", "Fator Geral")
    ]
    for item_name, key_name in process_total_items_for_pdf:
        value_brl = st.session_state.process_totals.get(key_name, "N/A")
        value_usd_key = key_name.replace(" (R$)", " (US$)").replace(" (KG)", "")
        value_usd = st.session_state.process_totals.get(value_usd_key, "N/A")
        process_totals_data.append([item_name, value_brl, value_usd])

    table_process_totals = Table(process_totals_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch])
    table_process_totals.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(table_process_totals)
    story.append(Spacer(1, 0.2*inch))

    # --- Impostos ---
    story.append(Paragraph("Impostos:", style_heading))
    impostos_data = [
        ["Imposto", "Valor"]
    ]
    for tax, value in st.session_state.taxes_data.items():
        impostos_data.append([tax.upper(), value])
    
    table_impostos = Table(impostos_data, colWidths=[2.5*inch, 2.5*inch])
    table_impostos.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(table_impostos)
    story.append(Spacer(1, 0.2*inch))

    # --- Despesas ---
    story.append(Paragraph("Despesas:", style_heading))
    despesas_data = [
        ["Item", "Valor"]
    ]
    for item, value in st.session_state.expenses_display.items():
        despesas_data.append([item.replace('_', ' ').title(), value])

    table_despesas = Table(despesas_data, colWidths=[2.5*inch, 2.5*inch])
    table_despesas.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(table_despesas)
    story.append(Spacer(1, 0.2*inch))

    # --- Contratos de Câmbio ---
    story.append(Paragraph("Contratos de Câmbio:", style_heading))
    cambio_data = [
        ["Nº Contrato", "Dólar", "Valor (USD)"]
    ]
    for index, row in st.session_state.contracts_df.iterrows():
        num_contrato = row['Nº Contrato']
        dolar = _format_float(row['Dólar'], 4)
        valor_usd = _format_float(row['Valor (US$)'], 2, prefix="US$ ")
        try:
            if (float(row['Dólar']) > 0 and float(row['Valor (US$)']) > 0) or (num_contrato and num_contrato != f"Contrato {index+1}"):
                cambio_data.append([num_contrato, dolar, valor_usd])
        except ValueError:
            pass
    
    cambio_data.append(["Soma Total (USD):", "", _format_float(soma_contratos_usd, 2, prefix='US$ ')])
    cambio_data.append(["Diferença (USD):", "", _format_float(diferenca_contratos_usd, 2, prefix='US$ ')])

    table_cambio = Table(cambio_data, colWidths=[1.5*inch, 1.5*inch, 1.5*inch])
    table_cambio.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (0,-1), 'LEFT'),
        ('ALIGN', (1,0), (-1,-1), 'RIGHT'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTNAME', (0,-2), (-1,-1), 'Helvetica-Bold'),
        ('ALIGN', (0,-2), (0,-1), 'LEFT'),
        ('ALIGN', (1,-2), (-1,-1), 'RIGHT'),
    ]))
    story.append(table_cambio)
    story.append(Spacer(1, 0.2*inch))

    # --- Mudar para o modo paisagem para Detalhes dos Itens ---
    story.append(PageBreak()) # Quebra de página antes de mudar o layout
    story.append(NextPageTemplate('LandscapePage')) # Solicita que a próxima página use o template paisagem
    story.append(Spacer(1, 0.1*inch)) # Adiciona um pequeno espaçador para garantir que a template seja aplicada

    story.append(Paragraph("Detalhes dos Itens:", style_heading))

    item_headers_pdf = [
        "Código", "NCM", "SKU", "Qtd", "CIF Unit.",
        "II", "IPI", "PIS", "COFINS", "Fator",
        "VLME (BRL)", "VLMD (BRL)"
    ]
    item_data_for_pdf = [item_headers_pdf]

    for index, row in itens_df_calculated.iterrows():
        if row["Código ERP"] == "TOTAL": # Skip the total row
            continue
        row_values_for_pdf = [
            row["Código ERP"], row["NCM"], row["SKU"], row["Quantidade"], row["CIF Unitário"],
            row["II %"], row["IPI %"], row["PIS %"], row["COFINS %"], row["Fator de Internação"],
            row["VLME (BRL)"], row["VLMD (BRL)"]
        ]
        item_data_for_pdf.append(row_values_for_pdf)

    # Definir larguras das colunas para modo paisagem e aumentar SKU
    col_widths_pdf = [
        0.5*inch,  # Código ERP
        0.6*inch,  # NCM
        2.0*inch,  # SKU (aumentado para 2.0 polegadas)
        0.4*inch,  # Qtd
        0.7*inch,  # CIF Unit.
        0.4*inch,  # II %
        0.4*inch,  # IPI %
        0.4*inch,  # PIS %
        0.4*inch,  # COFINS %
        0.4*inch,  # Fator Intern.
        0.9*inch,  # VLME (BRL)
        0.9*inch   # VLMD (BRL)
    ]
    
    table_items = Table(item_data_for_pdf, colWidths=col_widths_pdf)
    table_items.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.grey),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,-1), 'CENTER'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0,0), (-1,0), 6),
        ('BACKGROUND', (0,1), (-1,-1), colors.beige),
        ('GRID', (0,0), (-1,-1), 0.5, colors.black),
        ('FONTSIZE', (0,0), (-1,-1), 7),
        ('LEFTPADDING', (0,0), (-1,-1), 2),
        ('RIGHTPADDING', (0,0), (-1,-1), 2),
    ]))
    story.append(table_items)

    now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
    story.append(Paragraph(f"Gerado em: {now}", ParagraphStyle(name='Footer', parent=styles['Normal'], fontSize=8, alignment=TA_RIGHT, spaceBefore=12)))

    try:
        # Construir o documento com a lista de elementos
        doc.build(story)
        buffer.seek(0)
        return buffer, file_name
    except Exception as e:
        st.error(f"Erro ao gerar PDF: {str(e)}")
        return None, None
def _generate_cover_pdf(di_data, total_para_nf, process_totals, contracts_df):
    """Gera a capa do processo em PDF."""
    if not di_data:
        st.warning("Nenhum dado de DI carregado para gerar a capa.")
        return None, None

    referencia_processo = di_data[6] if di_data[6] else "SemReferencia"
    file_name = f"{referencia_processo}_Capa.pdf"

    try:
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4) # Capa permanece em A4 retrato
        
        # Lista para armazenar os elementos do PDF
        story = []
        styles = getSampleStyleSheet()
        
        style_center_bold_large = ParagraphStyle(name='CenterBoldLarge', parent=styles['Normal'],
                                             fontName='Helvetica-Bold', fontSize=16, alignment=TA_CENTER, spaceAfter=14)
        style_center_bold = ParagraphStyle(name='CenterBold', parent=styles['Normal'],
                                       fontName='Helvetica-Bold', fontSize=12, alignment=TA_CENTER, spaceAfter=8)
        style_normal = styles['Normal']
        style_normal.fontSize = 10
        style_normal.leading = 12
        style_value = ParagraphStyle(name='Value', parent=styles['Normal'],
                                 fontName='Helvetica', fontSize=10, alignment=TA_RIGHT)

        # Add logo image
        logo_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'assets', 'logo.png')
        if os.path.exists(logo_path):
            img = Image(logo_path)
            # Set image size - adjust width and height as needed
            img.drawWidth = 3*inch  
            img.drawHeight = 0.9*inch
            story.append(img)
        else:
            # Fallback to text if image not found
            story.append(Paragraph("PICHAU", style_center_bold_large))
        story.append(Spacer(1, 0.1*inch))

        story.append(Paragraph(f"REFERÊNCIA DO PROCESSO: {di_data[6] if di_data[6] else ''}", style_center_bold))
        story.append(Spacer(1, 0.1*inch))

        # Data from DI (index based on db_utils.get_declaracao_by_id)
        (id_db, numero_di, data_registro_db, valor_total_reais_xml,
         arquivo_origem, data_importacao, informacao_complementar,
         vmle_declaracao, frete_declaracao, seguro_declaracao, vmld_declaracao,
         ipi_total_declaracao, pis_pasep_total_declaracao, cofins_total_declaracao, icms_sc,
         taxa_cambial_usd_declaracao, taxa_siscomex_total_declaracao, numero_invoice,
         peso_bruto_total, peso_liquido_total, cnpj_importador, importador_nome,
         recinto, embalagem, quantidade_volumes_total, acrescimo_total_declaracao,
         imposto_importacao_total_declaracao, armazenagem_db, frete_nacional_db) = di_data

        vmle_usd_capa = vmle_declaracao / taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao > 0 else 0.0
        frete_usd_capa = frete_declaracao / taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao > 0 else 0.0
        seguro_usd_capa = seguro_declaracao / taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao > 0 else 0.0
        vmld_usd_capa = vmld_declaracao / taxa_cambial_usd_declaracao if taxa_cambial_usd_declaracao > 0 else 0.0

        # Tabela de Desembaraço
        desembaraco_data = [
            ["DI:", _format_di_number(numero_di)],
            ["DATA DI:", datetime.strptime(data_registro_db, "%Y-%m-%d").strftime("%d/%m/%Y") if data_registro_db else ""],
            ["DATA DESEMBARAÇO:", st.session_state.capa_data_desembaraco_var], # From session state
            ["CANAL:", st.session_state.capa_canal_var], # From session state
            ["TIPO DE IMPORTAÇÃO:", "DIRETA"], # Mock
        ]
        table_desembaraco = Table(desembaraco_data, colWidths=[2.5*inch, 2.5*inch])
        table_desembaraco.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        story.append(table_desembaraco)
        
        

        produtos_data = [
            ["FORNECEDOR:", st.session_state.capa_fornecedor_var],
            ["PRODUTOS:", st.session_state.capa_produtos_var],
            ["VOLUMES:", "CAIXA"], # Mock
            ["QTDE ITENS:", _format_int_no_float(sum(item[5] for item in st.session_state.itens_data if item[5] is not None))],
            ["QTDE VOLUMES:", _format_int(quantidade_volumes_total)],
        ]
        try: # Safely check if capa_quantidade_containers_var is a valid number
            if st.session_state.capa_modal_var == "MARITIMO" and float(st.session_state.capa_quantidade_containers_var.replace(',', '.')) > 0:
                produtos_data.append(["QUANTIDADE DE CONTAINERS:", st.session_state.capa_quantidade_containers_var])
        except ValueError:
            pass # If not a valid number, skip adding it

        produtos_data.append(["PESO BRUTO (kg):", _format_weight_no_kg(peso_bruto_total)])

        table_produtos = Table(produtos_data, colWidths=[2.5*inch, 2.5*inch])
        table_produtos.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        story.append(table_produtos)
        story.append(Spacer(1, 0.2*inch))

        story.append(Paragraph("INFORMAÇÕES GERAIS:", style_center_bold))
        info_gerais_data = [
            ["ORIGEM:", "SHENZHEN"], # Mock
            ["DESTINO:", "NAVEGANTES"], # Mock
            ["MODAL:", st.session_state.capa_modal_var],
            ["INCOTERM:", st.session_state.capa_incoterm_var],
        ]
        table_info_gerais = Table(info_gerais_data, colWidths=[2.5*inch, 2.5*inch])
        table_info_gerais.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        story.append(table_info_gerais)
        story.append(Spacer(1, 0.2*inch))

        story.append(Paragraph("VALORES (USD):", style_center_bold))
        valores_usd_data = [
            ["VMLE:", _format_float(vmle_usd_capa, 2, prefix="$ ")],
            ["FRETE:", _format_float(frete_usd_capa, 2, prefix="$ ")],
            ["SEGURO:", _format_float(seguro_usd_capa, 2, prefix="$ ")],
            ["VMLD:", _format_float(vmld_usd_capa, 2, prefix="$ ")],
            ["CÂMBIO:", _format_float(taxa_cambial_usd_declaracao, 4)],
        ]
        table_valores_usd = Table(valores_usd_data, colWidths=[2.5*inch, 2.5*inch])
        table_valores_usd.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        story.append(table_valores_usd)
        story.append(Spacer(1, 0.2*inch))

        story.append(Paragraph("NACIONAL:", style_center_bold))
        nacional_data = [
            ["TRANSPORTADORA:", st.session_state.capa_transportadora_var],
            ["NF ENTRADA:", st.session_state.capa_nf_entrada_var],
            ["TOTAL IMPORTAÇÃO (R$):", _format_currency(total_para_nf)],
            ["FATOR BRUTO:", process_totals.get("Fator Geral", "N/A")],
        ]
        table_nacional = Table(nacional_data, colWidths=[2.5*inch, 2.5*inch])
        table_nacional.setStyle(TableStyle([
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('FONTNAME', (0,0), (0,-1), 'Helvetica-Bold'),
            ('GRID', (0,0), (-1,-1), 0.5, colors.black),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
        ]))
        story.append(table_nacional)
        story.append(Spacer(1, 0.2*inch))

        now = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        story.append(Paragraph(f"Gerado em: {now}", ParagraphStyle(name='Footer', parent=styles['Normal'], fontSize=8, alignment=TA_RIGHT, spaceBefore=12)))

        # Constrói o documento com a lista story
        doc.build(story)  # Passa a lista story, não o buffer
        buffer.seek(0)
        return buffer, file_name
        
    except Exception as e:
        logger.exception("Erro ao gerar PDF da capa")
        st.error(f"Erro ao gerar PDF da capa: {str(e)}")
        return None, None

# --- Função para atualizar todos os cálculos na session_state ---
def update_all_calculations():
    """Recalcula todos os totais e atualiza a session_state."""
    if st.session_state.di_data:
        process_totals, taxes_data, expenses_display, itens_df_calculated, soma_contratos_usd, diferenca_contratos_usd = \
            perform_calculations(st.session_state.di_data, st.session_state.itens_data,
                                  st.session_state.expense_inputs, st.session_state.contracts_df)
        
        st.session_state.process_totals = process_totals
        st.session_state.taxes_data = taxes_data
        st.session_state.expenses_display = expenses_display
        st.session_state.soma_contratos_usd = soma_contratos_usd
        st.session_state.diferenca_contratos_usd = diferenca_contratos_usd
        
        # Atualiza total_para_nf
        st.session_state.total_para_nf = expenses_display.get("TOTAL PARA NF", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            st.session_state.total_para_nf = float(st.session_state.total_para_nf)
        except ValueError:
            st.session_state.total_para_nf = 0.0
    # Removido st.rerun() daqui, pois o Streamlit reexecuta naturalmente ao alterar session_state

# --- Função Principal da Página de Custo ---
def show_page():
    background_image_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'assets', 'logo_navio_atracado.png')
    set_background_image(background_image_path)
    # Inicializa as variáveis de estado no início da função
    if 'di_data' not in st.session_state:
        st.session_state.di_data = None
    if 'itens_data' not in st.session_state:
        st.session_state.itens_data = []
    if 'item_erp_codes' not in st.session_state:
        st.session_state.item_erp_codes = {}
    if 'expense_inputs' not in st.session_state:
        st.session_state.expense_inputs = {
            'afrmm': 0.00, 'siscoserv': 0.00, 'descarregamento': 0.00, 'taxas_destino': 0.00, 'multa': 0.00,
        }
    if 'contracts_df' not in st.session_state:
        st.session_state.contracts_df = pd.DataFrame({
            'Nº Contrato': [f"Contrato {i+1}" for i in range(10)], 'Dólar': [0.0000] * 10, 'Valor (US$)': [0.00] * 10
        })
    if 'custo_search_ref_input' not in st.session_state:
        st.session_state.custo_search_ref_input = "PCH-"
    if 'process_totals' not in st.session_state:
        st.session_state.process_totals = {}
    if 'taxes_data' not in st.session_state:
        st.session_state.taxes_data = {}
    if 'expenses_display' not in st.session_state:
        st.session_state.expenses_display = {}
    if 'soma_contratos_usd' not in st.session_state:
        st.session_state.soma_contratos_usd = 0.0
    if 'diferenca_contratos_usd' not in st.session_state:
        st.session_state.diferenca_contratos_usd = 0.0
    if 'total_para_nf' not in st.session_state:
        st.session_state.total_para_nf = 0.0
    # Inicializa variáveis para a capa no session_state (se ainda não estiverem)
    if 'capa_data_desembaraco_var' not in st.session_state:
        st.session_state.capa_data_desembaraco_var = datetime.now().strftime("%d/%m/%Y")
    if 'capa_canal_var' not in st.session_state:
        st.session_state.capa_canal_var = "VERDE"
    if 'capa_fornecedor_var' not in st.session_state:
        st.session_state.capa_fornecedor_var = "" # Será preenchido após carregar DI
    if 'capa_produtos_var' not in st.session_state:
        st.session_state.capa_produtos_var = ""
    if 'capa_modal_var' not in st.session_state:
        st.session_state.capa_modal_var = ""
    if 'capa_quantidade_containers_var' not in st.session_state:
        st.session_state.capa_quantidade_containers_var = "0"
    if 'capa_incoterm_var' not in st.session_state:
        st.session_state.capa_incoterm_var = ""
    if 'capa_transportadora_var' not in st.session_state:
        st.session_state.capa_transportadora_var = ""
    if 'capa_nf_entrada_var' not in st.session_state:
        st.session_state.capa_nf_entrada_var = ""
    
    # Adicionada flag para controlar a atualização do contracts_df
    if 'contracts_df_updated_by_button' not in st.session_state:
        st.session_state.contracts_df_updated_by_button = True # Começa como True para carregar os dados iniciais

    st.subheader("Processo")
    col1_search, col2_search, col3 = st.columns([0.4, 0.2, 0.4])

    with col1_search:
        search_ref = st.text_input("Referência do Processo", value=st.session_state.custo_search_ref_input, key="custo_search_ref_input_widget")
        st.session_state.custo_search_ref_input = search_ref # Atualiza o valor persistente

    with col2_search:
        st.write("") # Espaço para alinhar o botão
        st.write("")
        if st.button("Pesquisar", key="custo_search_button"):
            declaracao = get_declaracao_by_referencia(search_ref)
            if declaracao:
                st.session_state.di_data = declaracao
                st.session_state.itens_data = get_itens_by_declaracao_id(declaracao[0])
                
                # Load existing ERP codes from DB for items
                st.session_state.item_erp_codes = {}
                if st.session_state.itens_data:
                    for item_tuple in st.session_state.itens_data:
                        # Ensure item_tuple is a tuple/list, not sqlite3.Row
                        if isinstance(item_tuple, sqlite3.Row):
                            item_tuple = tuple(item_tuple)
                        item_id_db = item_tuple[0]
                        codigo_erp_from_db = item_tuple[17] # 18th element is codigo_erp_item
                        if codigo_erp_from_db:
                            st.session_state.item_erp_codes[item_id_db] = codigo_erp_from_db

                # Load existing expenses and contracts from DB
                expenses_db, contracts_db = get_process_cost_data(declaracao[0])
                if expenses_db:
                    st.session_state.expense_inputs = {
                        'afrmm': expenses_db[0],
                        'siscoserv': expenses_db[1],
                        'descarregamento': expenses_db[2],
                        'taxas_destino': expenses_db[3],
                        'multa': expenses_db[4],
                    }
                else: # Default if no data found
                    st.session_state.expense_inputs = {
                        'afrmm': 0.00, 'siscoserv': 0.00, 'descarregamento': 0.00, 'taxas_destino': 0.00, 'multa': 0.00,
                    }
                
                # Initialize contracts_df
                contracts_df_data = []
                if contracts_db:
                    for contract in contracts_db:
                        # Ensure contract is a tuple/list, not sqlite3.Row
                        if isinstance(contract, sqlite3.Row):
                            contract = tuple(contract)
                        contracts_df_data.append({
                            'Nº Contrato': contract[0],
                            'Dólar': contract[1],
                            'Valor (US$)': contract[2]
                        })
                else: # Default empty contracts
                    for i in range(10):
                        contracts_df_data.append({
                            'Nº Contrato': f"Contrato {i+1}",
                            'Dólar': 0.0000,
                            'Valor (US$)': 0.00
                        })
                st.session_state.contracts_df = pd.DataFrame(contracts_df_data)

                # Preenche o primeiro contrato com a taxa cambial da DI e o VMLE em Dólar se não houver contratos carregados
                # e também os demais campos de Dólar
                if declaracao[15] is not None and declaracao[15] > 0: # taxa_cambial_usd_declaracao
                    taxa_cambial = declaracao[15]
                    vmle_brl = declaracao[7] if declaracao[7] is not None else 0.0 # vmle_declaracao
                    if taxa_cambial > 0:
                        vmle_usd = vmle_brl / taxa_cambial
                        
                        # Se não há contratos carregados, preenche o primeiro
                        if not contracts_db:
                            st.session_state.contracts_df.loc[0, 'Dólar'] = taxa_cambial
                            st.session_state.contracts_df.loc[0, 'Valor (US$)'] = vmle_usd
                        
                        # Preenche todas as linhas da coluna 'Dólar' com a taxa cambial da DI
                        st.session_state.contracts_df['Dólar'] = taxa_cambial


                # Atualiza capa_fornecedor_var com o nome do importador
                st.session_state.capa_fornecedor_var = declaracao[21] if declaracao[21] else ""


                st.success(f"Dados do processo '{search_ref}' carregados!")
                st.session_state.contracts_df_updated_by_button = True # Força a atualização dos cálculos
                
            else:
                st.session_state.di_data = None
                st.session_state.itens_data = []
                st.session_state.item_erp_codes = {}
                st.session_state.expense_inputs = {'afrmm': 0.00, 'siscoserv': 0.00, 'descarregamento': 0.00, 'taxas_destino': 0.00, 'multa': 0.00}
                st.session_state.contracts_df = pd.DataFrame({'Nº Contrato': [f"Contrato {i+1}" for i in range(10)], 'Dólar': [0.0000] * 10, 'Valor (US$)': [0.00] * 10})
                st.session_state.capa_fornecedor_var = "" # Limpa o fornecedor da capa
                st.warning(f"Nenhum processo encontrado com a referência: {search_ref}")
            
            # Após carregar os dados, force um rerun para que os cálculos iniciais sejam feitos
            update_all_calculations()


    if st.session_state.di_data:
        st.markdown(f"**Processo:** {st.session_state.di_data[6]}")
    else:
        st.markdown("**Processo:** N/A")

    st.markdown("---")

    # --- Abas para Totais, Impostos, Despesas e Contratos ---
    tab1, tab2, tab3, tab4, tab5 = st.tabs(["Total do Processo", "Impostos", "Despesas", "Contratos de Câmbio", "Comparativos"])

    # Realiza os cálculos
    # Os cálculos agora são feitos uma vez no início da página e atualizados pelos callbacks
    # ou pelo botão de atualização dos contratos
    
    # Se a flag contracts_df_updated_by_button for True, recalcula e depois seta para False
    if st.session_state.contracts_df_updated_by_button:
        process_totals, taxes_data, expenses_display, itens_df_calculated, soma_contratos_usd, diferenca_contratos_usd = \
            perform_calculations(st.session_state.di_data, st.session_state.itens_data, st.session_state.expense_inputs, st.session_state.contracts_df)
        
        st.session_state.process_totals = process_totals
        st.session_state.taxes_data = taxes_data
        st.session_state.expenses_display = expenses_display
        st.session_state.soma_contratos_usd = soma_contratos_usd
        st.session_state.diferenca_contratos_usd = diferenca_contratos_usd
        
        # Atualiza total_para_nf
        st.session_state.total_para_nf = expenses_display.get("TOTAL PARA NF", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip()
        try:
            st.session_state.total_para_nf = float(st.session_state.total_para_nf)
        except ValueError:
            st.session_state.total_para_nf = 0.0
        
        st.session_state.contracts_df_updated_by_button = False # Reseta a flag

    # Usa os valores armazenados no session_state para exibir
    process_totals = st.session_state.process_totals
    taxes_data = st.session_state.taxes_data
    expenses_display = st.session_state.expenses_display
    soma_contratos_usd = st.session_state.soma_contratos_usd
    diferenca_contratos_usd = st.session_state.diferenca_contratos_usd
    itens_df_calculated = perform_calculations(st.session_state.di_data, st.session_state.itens_data, st.session_state.expense_inputs, st.session_state.contracts_df)[3]


    with tab1:
        st.subheader("TOTAL DO PROCESSO")
        if process_totals:
            st.markdown("##### Valores do Processo")
            cols = st.columns(5)
            cols[0].markdown("**Item**")
            cols[1].markdown("**Valor (R$)**")
            cols[2].markdown("**Valor (US$)**")

            items_to_display = [
                ("Taxa Cambial", process_totals["Taxa Cambial"], "--"),
                ("VMLE", process_totals["VMLE (R$)"], process_totals["VMLE (US$)"]),
                ("Frete", process_totals["Frete (R$)"], process_totals["Frete (US$)"]),
                ("Seguro", process_totals["Seguro (R$)"], process_totals["Seguro (US$)"]),
                ("VMLD (CIF)", process_totals["VMLD (CIF) (R$)"], process_totals["VMLD (CIF) (US$)"]),
                ("Acréscimo", process_totals["Acréscimo (R$)"], process_totals["Acréscimo (US$)"]),
                ("Peso Total (KG)", process_totals["Peso Total (KG)"], "--"),
                ("SISCOMEX", process_totals["SISCOMEX"], "--"),
                ("Despesas Operacionais", process_totals["Despesas Operacionais"], "--"),
                ("Fator Geral", process_totals["Fator Geral"], "--")
            ]
            for item, val_brl, val_usd in items_to_display:
                cols = st.columns(5)
                cols[0].write(item)
                cols[1].write(val_brl)
                cols[2].write(val_usd)
        else:
            st.info("Carregue os dados da DI para ver os totais do processo.")

    with tab2:
        
        if taxes_data:
            st.markdown("##### Impostos Totais")
            cols = st.columns(5)
            cols[0].markdown("**Imposto**")
            cols[1].markdown("**Valor**")
            for tax, value in taxes_data.items():
                cols = st.columns(5)
                cols[0].write(tax)
                cols[1].write(value)
        else:
            st.info("Carregue os dados da DI para ver os impostos.")

    with tab3:
        st.subheader("DESPESAS")
        if expenses_display:
            st.markdown("##### Despesas do Processo")
            cols = st.columns(4)
            with cols[0]:
                # Campos editáveis com on_change para atualizar os cálculos
                st.session_state.expense_inputs['afrmm'] = st.number_input("AFRMM", value=st.session_state.expense_inputs['afrmm'], format="%.2f", key="afrmm_input", on_change=update_all_calculations)
                st.session_state.expense_inputs['siscoserv'] = st.number_input("SISCOSERV", value=st.session_state.expense_inputs['siscoserv'], format="%.2f", key="siscoserv_input", on_change=update_all_calculations)
                st.session_state.expense_inputs['descarregamento'] = st.number_input("DESCARREGAMENTO", value=st.session_state.expense_inputs['descarregamento'], format="%.2f", key="descarregamento_input", on_change=update_all_calculations)
                st.session_state.expense_inputs['taxas_destino'] = st.number_input("TAXAS DESTINO", value=st.session_state.expense_inputs['taxas_destino'], format="%.2f", key="taxas_destino_input", on_change=update_all_calculations)
                st.session_state.expense_inputs['multa'] = st.number_input("MULTA", value=st.session_state.expense_inputs['multa'], format="%.2f", key="multa_input", on_change=update_all_calculations)

            st.markdown("---")
            st.markdown("##### Resumo das Despesas")
            cols = st.columns(2)
            cols[0].markdown("**Item**")
            cols[1].markdown("**Valor**")
            for item, value in expenses_display.items():
                cols = st.columns(2)
                cols[0].write(item)
                cols[1].write(value)

            if st.button("Salvar Despesas e Contratos", key="save_expenses_contracts_button"):
                if st.session_state.di_data:
                    declaracao_id = st.session_state.di_data[0] # ID da DI
                    success = save_process_cost_data(
                        declaracao_id,
                        st.session_state.expense_inputs['afrmm'],
                        st.session_state.expense_inputs['siscoserv'],
                        st.session_state.expense_inputs['descarregamento'],
                        st.session_state.expense_inputs['taxas_destino'],
                        st.session_state.expense_inputs['multa'],
                        st.session_state.contracts_df
                    )
                    if success:
                        st.success("Despesas e Contratos salvos com sucesso no banco de dados!")
                    else:
                        st.error("Falha ao salvar Despesas e Contratos.")
                else:
                    st.warning("Carregue um processo antes de tentar salvar despesas e contratos.")
        else:
            st.info("Carregue os dados da DI para ver e editar as despesas.")

    with tab4:
        st.subheader("CONTRATOS DE CÂMBIO")
        if st.session_state.di_data:
            st.markdown("##### Edite os Contratos de Câmbio")
            
            # Editor de contratos sem on_change para atualizar apenas com o botão
            # É crucial que contracts_df_temp seja inicializado com st.session_state.contracts_df
            # para que o editor comece com os valores atuais.
            col_1, col_2, col_3 = st.columns([0.5, 0.1, 0.2]) # Colunas para o editor e os totais
            with col_1:
                contracts_df_temp = st.data_editor(
                    st.session_state.contracts_df,
                    column_config={
                        "Nº Contrato": st.column_config.TextColumn("Nº Contrato", width="small"),
                        "Dólar": st.column_config.NumberColumn("Dólar", format="%.4f",width="small"),
                        "Valor (US$)": st.column_config.NumberColumn("Valor (US$)", format="$%.2f",width="medium"),
                    },
                    num_rows="dynamic",
                    hide_index=True,
                    use_container_width=True,
                    key="contracts_editor_no_live_update", # Nova chave para o editor
                )

            # Botão para atualizar os contratos e recalcular a página
            
            
            with col_3:
                if st.button("Atualizar Contratos", key="update_contracts_button"):
                    st.session_state.contracts_df = contracts_df_temp # Atualiza o DataFrame persistente
                    st.session_state.contracts_df_updated_by_button = True # Força o recálculo
                    st.rerun() # Dispara um rerun para que os cálculos reflitam as mudanças
                
                st.markdown("##### Totais dos Contratos")
                st.markdown(f"Soma Contratos (US$): {_format_float(soma_contratos_usd, 2, prefix='USD ')}")
                st.markdown(f"Diferença (US$): {_format_float(diferenca_contratos_usd, 2, prefix='USD ')}")
        else:
            st.info("Carregue os dados da DI para ver e editar os contratos.")

    with tab5:
        st.subheader("COMPARATIVOS")
        if st.session_state.di_data:
            # Desempacota os dados da DI novamente para fácil acesso aos valores do banco
            (id_db, numero_di, data_registro_db, valor_total_reais_xml,
             arquivo_origem, data_importacao, informacao_complementar,
             vmle_declaracao, frete_declaracao_db, seguro_declaracao_db, vmld_declaracao,
             ipi_total_declaracao_db, pis_pasep_total_declaracao_db, cofins_total_declaracao_db, icms_sc,
             taxa_cambial_usd_declaracao, taxa_siscomex_total_declaracao, numero_invoice,
             peso_bruto_total, peso_liquido_total, cnpj_importador, importador_importador_nome,
             recinto, embalagem, quantidade_volumes_total, acrescimo_total_declaracao,
             imposto_importacao_total_declaracao_db, armazenagem_db, frete_nacional_db) = st.session_state.di_data

            st.markdown("##### Comparativo de Valores (Calculado vs. Declaração de Importação)")
            
            comparative_data = []

            # II
            ii_calculado = float(st.session_state.taxes_data.get("II", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            ii_banco = imposto_importacao_total_declaracao_db if imposto_importacao_total_declaracao_db is not None else 0.0
            comparative_data.append(["II", _format_currency(ii_calculado), _format_currency(ii_banco), _format_currency(ii_calculado - ii_banco)])

            # IPI
            ipi_calculado = float(st.session_state.taxes_data.get("IPI", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            ipi_banco = ipi_total_declaracao_db if ipi_total_declaracao_db is not None else 0.0
            comparative_data.append(["IPI", _format_currency(ipi_calculado), _format_currency(ipi_banco), _format_currency(ipi_calculado - ipi_banco)])

            # PIS
            pis_calculado = float(st.session_state.taxes_data.get("PIS", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            pis_banco = pis_pasep_total_declaracao_db if pis_pasep_total_declaracao_db is not None else 0.0
            comparative_data.append(["PIS", _format_currency(pis_calculado), _format_currency(pis_banco), _format_currency(pis_calculado - pis_banco)])

            # COFINS
            cofins_calculado = float(st.session_state.taxes_data.get("COFINS", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            cofins_banco = cofins_total_declaracao_db if cofins_total_declaracao_db is not None else 0.0
            comparative_data.append(["COFINS", _format_currency(cofins_calculado), _format_currency(cofins_banco), _format_currency(cofins_calculado - cofins_banco)])

            # FRETE
            frete_calculado = float(st.session_state.process_totals.get("Frete (R$)", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            frete_banco = frete_declaracao_db if frete_declaracao_db is not None else 0.0
            comparative_data.append(["FRETE", _format_currency(frete_calculado), _format_currency(frete_banco), _format_currency(frete_calculado - frete_banco)])

            # SEGURO
            seguro_calculado = float(st.session_state.process_totals.get("Seguro (R$)", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            seguro_banco = seguro_declaracao_db if seguro_declaracao_db is not None else 0.0
            comparative_data.append(["SEGURO", _format_currency(seguro_calculado), _format_currency(seguro_banco), _format_currency(seguro_calculado - seguro_banco)])

            # DESPESAS OPERACIONAIS (Total de Despesas Operacionais)
            # Calculado a partir da soma das despesas editáveis e fixas
            despesas_operacionais_calculado = float(st.session_state.expenses_display.get("TOTAL", "R$ 0,00").replace('R$', '').replace('.', '').replace(',', '.').strip())
            
            # Para o valor do banco para "Despesas Operacionais", você precisa ter um campo no banco que represente
            # essa soma. Se não houver, ou se for a soma de campos específicos do banco (armazenagem, frete_nacional, siscomex)
            # mais os valores fixos que não vêm do banco, precisa ser explicitado.
            # No seu código atual, total_despesas_operacionais_db parece estar tentando somar alguns desses.
            # Vou mantê-lo assim, assumindo que ele reflete a "parte do banco" das despesas operacionais.
            total_despesas_operacionais_db_for_comparison = (armazenagem_db if armazenagem_db is not None else 0.0) + \
                                                             (frete_nacional_db if frete_nacional_db is not None else 0.0) + \
                                                             (taxa_siscomex_total_declaracao if taxa_siscomex_total_declaracao is not None else 0.0) + \
                                                             1000.00 + 0.00 # Honorário Despachante Fixo e Envio Docs Fixo

            # Adiciona as despesas editáveis do input, pois elas são persistidas e podem vir do banco
            # (afrmm, siscoserv, descarregamento, taxas_destino, multa)
            total_despesas_operacionais_db_for_comparison += st.session_state.expense_inputs['afrmm']
            total_despesas_operacionais_db_for_comparison += st.session_state.expense_inputs['siscoserv']
            total_despesas_operacionais_db_for_comparison += st.session_state.expense_inputs['descarregamento']
            total_despesas_operacionais_db_for_comparison += st.session_state.expense_inputs['taxas_destino']
            total_despesas_operacionais_db_for_comparison += st.session_state.expense_inputs['multa']
            
            comparative_data.append(["DESPESAS OPERACIONAIS", _format_currency(despesas_operacionais_calculado), _format_currency(total_despesas_operacionais_db_for_comparison), _format_currency(despesas_operacionais_calculado - total_despesas_operacionais_db_for_comparison)])


            df_comparativo = pd.DataFrame(comparative_data, columns=["Item", "Valor Calculado", "Valor do Banco", "Diferença"])
            st.dataframe(df_comparativo, hide_index=True, use_container_width=True)

        else:
            st.info("Carregue os dados da DI para ver os comparativos.")

    st.markdown("---")

    # --- Detalhes dos Itens (Treeview) ---
    st.subheader("DETALHES DO ITEM")

    if st.session_state.di_data:
        col_buttons_item = st.columns(4)
        with col_buttons_item[0]:
            excel_buffer, excel_filename = _generate_excel_for_cadastro(st.session_state.di_data, st.session_state.itens_data, st.session_state.item_erp_codes)
            if excel_buffer:
                st.download_button(
                    label="Excel para Cadastro",
                    data=excel_buffer,
                    file_name=excel_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="download_excel_cadastro"
                )
        with col_buttons_item[1]:
            uploaded_file = st.file_uploader("Importar Código ERP do Excel", type=["xlsx"], key="upload_erp_excel")
            if uploaded_file is not None:
                updates_count = _import_excel_for_cadastro(uploaded_file, st.session_state.itens_data)
                if updates_count > 0:
                    st.success(f"{updates_count} Códigos ERP atualizados com sucesso!")
                    st.rerun() # Recarrega a página para refletir as mudanças
                else:
                    st.warning("Nenhum Código ERP foi atualizado.")

        

        st.markdown("---")

        # Tabela de Itens
        st.dataframe(
            itens_df_calculated.drop(columns=["ID"]), # Remove a coluna ID para exibição
            hide_index=True,
            use_container_width=True,
            height=400,
            column_config={
                "Código ERP": st.column_config.TextColumn("Código ERP", width="small"),
                "NCM": st.column_config.TextColumn("NCM", width="small"),
                "SKU": st.column_config.TextColumn("SKU", width="medium"),
                "Descrição": st.column_config.TextColumn("Descrição", width="large"),
                "Quantidade": st.column_config.NumberColumn("Quantidade", format="%.0f", width="small"), # Formato para inteiro
                "Peso Unitário": st.column_config.TextColumn("Peso Unitário", width="small"),
                "CIF Unitário": st.column_config.TextColumn("CIF Unitário", width="small"),
                "VLME (BRL)": st.column_config.TextColumn("VLME (BRL)", width="small"),
                "VLMD (BRL)": st.column_config.TextColumn("VLMD (BRL)", width="small"),
                "II (BRL)": st.column_config.TextColumn("II (BRL)", width="small"),
                "IPI (BRL)": st.column_config.TextColumn("IPI (BRL)", width="small"),
                "PIS (BRL)": st.column_config.TextColumn("PIS (BRL)", width="small"),
                "COFINS (BRL)": st.column_config.TextColumn("COFINS (BRL)", width="small"),
                "II %": st.column_config.TextColumn("II %", width="small"),
                "IPI %": st.column_config.TextColumn("IPI %", width="small"),
                "PIS %": st.column_config.TextColumn("PIS %", width="small"),
                "COFINS %": st.column_config.TextColumn("COFINS %", width="small"),
                "ICMS %": st.column_config.TextColumn("ICMS %", width="small"),
                "Frete R$": st.column_config.TextColumn("Frete R$", width="small"),
                "Seguro R$": st.column_config.TextColumn("Seguro R$", width="small"),
                "Unitário US$ DI": st.column_config.TextColumn("Unitário US$ DI", width="small"),
                "Despesas Rateada": st.column_config.TextColumn("Despesas Rateada", width="small"),
                "Total de Despesas": st.column_config.TextColumn("Total de Despesas", width="small"),
                "Total Unitário": st.column_config.TextColumn("Total Unitário", width="small"),
                "Variação Cambial": st.column_config.TextColumn("Variação Cambial", width="small"),
                "Total Unitário com Variação": st.column_config.TextColumn("Total Unitário com Variação", width="small"),
                "Fator de Internação": st.column_config.TextColumn("Fator de Internação", width="small"),
                "Fator por Adição": st.column_config.TextColumn("Fator por Adição", width="small"),
            }
        )
        col_buttons_relatorio = st.columns(2)
        with col_buttons_relatorio[0]:
            
            with st.expander("Editar Capa para PDF"):
                    st.text_input("Data Desembaraço", value=st.session_state.capa_data_desembaraco_var, key="capa_data_desembaraco_input")
                    st.session_state.capa_data_desembaraco_var = st.session_state.capa_data_desembaraco_input

                    st.selectbox("Canal", options=["VERDE", "AMARELO", "VERMELHO", "CINZA"], index=["VERDE", "AMARELO", "VERMELHO", "CINZA"].index(st.session_state.capa_canal_var) if st.session_state.capa_canal_var in ["VERDE", "AMARELO", "VERMELHO", "CINZA"] else 0, key="capa_canal_input")
                    st.session_state.capa_canal_var = st.session_state.capa_canal_input

                    st.text_input("Fornecedor", value=st.session_state.capa_fornecedor_var, key="capa_fornecedor_input")
                    st.session_state.capa_fornecedor_var = st.session_state.capa_fornecedor_input

                    st.text_area("Produtos", value=st.session_state.capa_produtos_var, key="capa_produtos_input")
                    st.session_state.capa_produtos_var = st.session_state.capa_produtos_input

                    st.selectbox("Modal", options=["", "AEREO", "MARITIMO"], index=["", "AEREO", "MARITIMO"].index(st.session_state.capa_modal_var) if st.session_state.capa_modal_var in ["", "AEREO", "MARITIMO"] else 0, key="capa_modal_input")
                    st.session_state.capa_modal_var = st.session_state.capa_modal_input

                    if st.session_state.capa_modal_var == "MARITIMO":
                        st.text_input("Quantidade de Containers", value=st.session_state.capa_quantidade_containers_var, key="capa_containers_input")
                        st.session_state.capa_quantidade_containers_var = st.session_state.capa_containers_input
                    else:
                        st.session_state.capa_quantidade_containers_var = "0" # Reset if not maritime

                    st.selectbox("Incoterm", options=["", "EXW","FCA","FAS","FOB","CFR","CIF","CPT","CIP","DPU","DAP","DDP"], index=["", "EXW","FCA","FAS","FOB","CFR","CIF","CPT","CIP","DPU","DAP","DDP"].index(st.session_state.capa_incoterm_var) if st.session_state.capa_incoterm_var in ["", "EXW","FCA","FAS","FOB","CFR","CIF","CPT","CIP","DPU","DAP","DDP"] else 0, key="capa_incoterm_input")
                    st.session_state.capa_incoterm_var = st.session_state.capa_incoterm_input

                    st.text_input("Transportadora", value=st.session_state.capa_transportadora_var, key="capa_transportadora_input")
                    st.session_state.capa_transportadora_var = st.session_state.capa_transportadora_input

                    st.text_input("NF Entrada", value=st.session_state.capa_nf_entrada_var, key="capa_nf_entrada_input")
                    st.session_state.capa_nf_entrada_var = st.session_state.capa_nf_entrada_input

                    pdf_cover_buffer, pdf_cover_filename = _generate_cover_pdf(st.session_state.di_data, st.session_state.total_para_nf, st.session_state.process_totals, st.session_state.contracts_df)
                    if pdf_cover_buffer:
                        st.download_button(
                            label="Gerar Capa PDF",
                            data=pdf_cover_buffer,
                            file_name=pdf_cover_filename,
                            mime="application/pdf",
                            key="download_cover_pdf"
                        )
            pdf_buffer, pdf_filename = _generate_process_report_pdf(st.session_state.di_data, itens_df_calculated, soma_contratos_usd, diferenca_contratos_usd)
            if pdf_buffer:
                st.download_button(
                    label="Imprimir Relatório (PDF)",
                    data=pdf_buffer,
                    file_name=pdf_filename,
                    mime="application/pdf",
                    key="download_report_pdf"
                )
            
        st.markdown("---")        # Popup para edição da capa            
